"""
Prepara a participacao da cultura no orcamento da Uniao (2011-2025) para o
grafico de linhas, cruzando duas bases ja processadas:

- eixo1/orcamento/data/processed/federal/federal_final.csv, o gasto federal
  pleno em cultura por fonte (a mesma base de `federal-por-fonte.json`);
- eixo1/orcamento/data/processed/federal/RCL_2011_2025_Uniao_Resumido.xlsx, a
  Receita Corrente Liquida da Uniao apurada no RREO do Tesouro.

A conta e a mesma que `eixo1/orcamento/scripts/federal/gasto federal.R` faz em
`perc_rcl`: gasto pleno sobre RCL, ano a ano. A serie comeca em 2011 porque e
onde comeca a planilha de RCL, nao porque o gasto comece ali — a serie de gasto
vai a 2003.

Os dois lados vem em valores NOMINAIS, e e isso que a conta pede: a RCL e
apurada em reais correntes de cada ano, e deflacionar so um dos lados
inventaria uma participacao que ninguem mediu. Como o deflator multiplicaria
numerador e denominador do mesmo ano pelo mesmo indice, a participacao e
identica nas duas medidas — corrigir os dois lados seria trabalho para chegar
ao mesmo numero.

Uso:
    python3 scripts/build_participacao_rcl.py

Requer openpyxl (a RCL so existe em .xlsx).
"""

import csv
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

PROJECT_DIR = Path(__file__).resolve().parent.parent
REPO_ROOT = PROJECT_DIR.parent
FEDERAL_CSV = REPO_ROOT / "eixo1/orcamento/data/processed/federal/federal_final.csv"
RCL_XLSX = (
    REPO_ROOT
    / "eixo1/orcamento/data/processed/federal/RCL_2011_2025_Uniao_Resumido.xlsx"
)
OUT_PATH = PROJECT_DIR / "src/data/participacao-rcl.json"

# As duas fontes do gasto pleno que nao sao despesa orcamentaria: sao imposto
# que o Tesouro deixou de arrecadar para que o contribuinte o aplicasse em
# projeto cultural. Entram no numerador porque a figura mede o esforco federal
# em cultura por inteiro, e saem separadas porque so a outra parte disputa o
# orcamento com as demais politicas.
RENUNCIA = {"Lei Rouanet", "Incentivo (ANCINE)"}


def ler_gasto() -> dict[int, dict[str, float]]:
    """{ano: {pleno, direto, renuncia}} em reais nominais."""
    por_ano: dict[int, dict[str, float]] = defaultdict(
        lambda: {"pleno": 0.0, "direto": 0.0, "renuncia": 0.0}
    )

    with FEDERAL_CSV.open(encoding="utf-8") as f:
        for linha in csv.DictReader(f):
            ano = int(linha["ano"])
            valor = float(linha["valor_nominal"])
            registro = por_ano[ano]
            registro["pleno"] += valor
            registro["renuncia" if linha["fonte"] in RENUNCIA else "direto"] += valor

    return por_ano


def ler_rcl() -> dict[int, float]:
    """{ano: RCL} em reais nominais. A planilha ja traz a coluna em reais."""
    planilha = load_workbook(RCL_XLSX, read_only=True, data_only=True)["Planilha1"]
    linhas = planilha.iter_rows(values_only=True)
    cabecalho = list(next(linhas))
    i_ano, i_rcl = cabecalho.index("Ano"), cabecalho.index("RCL")

    rcl = {}
    for linha in linhas:
        if linha[i_ano] is None:
            continue
        rcl[int(linha[i_ano])] = float(linha[i_rcl])

    return rcl


def build() -> dict:
    gasto = ler_gasto()
    rcl = ler_rcl()

    # a serie e a intersecao: a RCL manda no inicio (2011) e o gasto, no fim
    anos = sorted(set(gasto) & set(rcl))
    if anos != list(range(anos[0], anos[-1] + 1)):
        raise ValueError(f"buraco na serie: {anos}")

    totais = [
        {
            "ano": ano,
            "pleno": round(gasto[ano]["pleno"], 2),
            "direto": round(gasto[ano]["direto"], 2),
            "renuncia": round(gasto[ano]["renuncia"], 2),
            "rcl": round(rcl[ano], 2),
        }
        for ano in anos
    ]

    def serie(key: str, label: str, campo: str) -> dict:
        return {
            "key": key,
            "label": label,
            "pontos": [
                {"ano": t["ano"], "pct": round(100 * t[campo] / t["rcl"], 4)}
                for t in totais
            ],
        }

    return {
        "unidade": "% da Receita Corrente Líquida da União",
        "denominador": "Receita Corrente Líquida da União",
        "totais": totais,
        # a figura desenha a primeira; a segunda fica disponivel para quem
        # quiser separar o que e despesa do que e imposto nao arrecadado
        "series": [
            serie("pleno", "Gasto federal pleno em cultura", "pleno"),
            serie("direto", "Execução orçamentária direta", "direto"),
        ],
    }


def main() -> None:
    dados = build()
    OUT_PATH.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    pontos = dados["series"][0]["pontos"]
    print(f"{OUT_PATH.relative_to(PROJECT_DIR)}: {len(pontos)} anos")
    print(f"  {pontos[0]['ano']}: {pontos[0]['pct']}%  ->  {pontos[-1]['ano']}: {pontos[-1]['pct']}%")


if __name__ == "__main__":
    main()
